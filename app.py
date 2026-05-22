import streamlit as st
import pandas as pd
import time
import re
import json
import io
from serpapi import GoogleSearch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  US STATE ABBREVIATIONS
# ─────────────────────────────────────────────
US_STATES = {
    "al": "Alabama", "ak": "Alaska", "az": "Arizona", "ar": "Arkansas",
    "ca": "California", "co": "Colorado", "ct": "Connecticut", "de": "Delaware",
    "fl": "Florida", "ga": "Georgia", "hi": "Hawaii", "id": "Idaho",
    "il": "Illinois", "in": "Indiana", "ia": "Iowa", "ks": "Kansas",
    "ky": "Kentucky", "la": "Louisiana", "me": "Maine", "md": "Maryland",
    "ma": "Massachusetts", "mi": "Michigan", "mn": "Minnesota", "ms": "Mississippi",
    "mo": "Missouri", "mt": "Montana", "ne": "Nebraska", "nv": "Nevada",
    "nh": "New Hampshire", "nj": "New Jersey", "nm": "New Mexico", "ny": "New York",
    "nc": "North Carolina", "nd": "North Dakota", "oh": "Ohio", "ok": "Oklahoma",
    "or": "Oregon", "pa": "Pennsylvania", "ri": "Rhode Island", "sc": "South Carolina",
    "sd": "South Dakota", "tn": "Tennessee", "tx": "Texas", "ut": "Utah",
    "vt": "Vermont", "va": "Virginia", "wa": "Washington", "wv": "West Virginia",
    "wi": "Wisconsin", "wy": "Wyoming", "dc": "District of Columbia",
}

# ─────────────────────────────────────────────
#  LOAD GOOGLE ADS LOCATIONS
# ─────────────────────────────────────────────
@st.cache_data
def load_locations():
    try:
        with open("locations.json", "r", encoding="utf-8") as f:
            data = json.load(f)
        # Flatten to a list of canonical strings like "Seattle, Washington, United States"
        locations = []
        for item in data:
            name = item.get("Name", "")
            if name:
                locations.append(name)
        return locations
    except Exception as e:
        st.error(f"locations.json load error: {e}")
        return []

# ─────────────────────────────────────────────
#  SMART LOCATION MATCHER  (FIXED)
#
#  KEY FIX: City is the primary signal.
#  State is only used as a tiebreaker when
#  multiple cities with the same name exist.
#  We NEVER force a wrong state onto a city.
# ─────────────────────────────────────────────
def find_precise_location(city: str, state: str, all_locs: list, country_name: str) -> str:
    city = str(city).strip()
    state_input = str(state).strip().lower()

    # Resolve state abbreviation → full name
    full_state_name = US_STATES.get(state_input, str(state).strip().title())

    # ── PASS 1: Exact match  city + state + country ──────────────────────
    # e.g. "Seattle, Washington, United States"
    for loc in all_locs:
        loc_lower = loc.lower()
        if (city.lower() in loc_lower
                and full_state_name.lower() in loc_lower
                and country_name.lower() in loc_lower):
            return loc

    # ── PASS 2: City + country (ignore state) ────────────────────────────
    # This is the CRITICAL fix:
    # If "Bothell, Florida, United States" doesn't exist but
    # "Bothell, Washington, United States" does, we still return it.
    country_city_matches = [
        loc for loc in all_locs
        if city.lower() in loc.lower()
        and country_name.lower() in loc.lower()
    ]

    if len(country_city_matches) == 1:
        # Only one city with this name in the country → safe to return
        return country_city_matches[0]

    if len(country_city_matches) > 1:
        # Multiple cities share this name → try state as tiebreaker
        state_filtered = [
            c for c in country_city_matches
            if full_state_name.lower() in c.lower()
        ]
        if state_filtered:
            return state_filtered[0]
        # State tiebreaker also failed → return first country match
        # (still better than a wrong-country result)
        return country_city_matches[0]

    # ── PASS 3: City only (any country) ──────────────────────────────────
    city_only_matches = [
        loc for loc in all_locs
        if loc.lower().startswith(city.lower())
    ]
    if city_only_matches:
        return city_only_matches[0]

    # ── PASS 4: Canonical fallback ────────────────────────────────────────
    # Google Ads accepts "City, State, Country" even if not in our list
    if full_state_name:
        return f"{city}, {full_state_name}, {country_name}"
    return f"{city}, {country_name}"


# ─────────────────────────────────────────────
#  SERPAPI: CHECK ORGANIC RANK
# ─────────────────────────────────────────────
def check_organic_rank(keyword: str, domain: str, location: str, api_key: str):
    """Returns (rank, found_url) or ("Not in Top 100", "-")"""
    try:
        params = {
            "engine": "google",
            "q": keyword,
            "location": location,
            "num": 100,
            "api_key": api_key,
            "gl": "us",
            "hl": "en",
        }
        search = GoogleSearch(params)
        results = search.get_dict()

        organic = results.get("organic_results", [])
        for i, result in enumerate(organic, start=1):
            url = result.get("link", "")
            if domain.lower().replace("www.", "") in url.lower().replace("www.", ""):
                return i, url

        return "Not in Top 100", "-"

    except Exception as e:
        return f"Error: {e}", "-"


# ─────────────────────────────────────────────
#  SERPAPI: CHECK MAPS / LOCAL PACK RANK
# ─────────────────────────────────────────────
def check_maps_rank(keyword: str, domain: str, location: str, api_key: str):
    """Returns rank in local pack or 'Not in Map Pack'"""
    try:
        params = {
            "engine": "google",
            "q": keyword,
            "location": location,
            "api_key": api_key,
            "gl": "us",
            "hl": "en",
        }
        search = GoogleSearch(params)
        results = search.get_dict()

        local_results = results.get("local_results", {})
        places = local_results.get("places", []) if isinstance(local_results, dict) else []

        for i, place in enumerate(places, start=1):
            website = place.get("website", "")
            if domain.lower().replace("www.", "") in website.lower().replace("www.", ""):
                return i

        return "Not in Map Pack"

    except Exception as e:
        return f"Error: {e}"


# ─────────────────────────────────────────────
#  EXCEL REPORT BUILDER
# ─────────────────────────────────────────────
def build_excel_report(results_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking Report"

    # ── Header styling ──
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    columns = ["#", "Keyword", "City", "State", "Targeted Location",
               "Organic Rank", "Maps Rank", "Found URL"]

    col_widths = [5, 40, 15, 8, 35, 14, 12, 50]

    for col_idx, (col_name, width) in enumerate(zip(columns, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # ── Data rows ──
    for row_idx, row in results_df.iterrows():
        excel_row = row_idx + 2
        values = [
            row_idx + 1,
            row.get("Keyword", ""),
            row.get("City", ""),
            row.get("State", ""),
            row.get("Targeted Location", ""),
            row.get("Organic Rank", ""),
            row.get("Maps Rank", ""),
            row.get("Found URL", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 8))

            # Color-code organic rank
            if col_idx == 6:
                if isinstance(value, int) and value <= 3:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                    cell.font = Font(color="276221", bold=True)
                elif isinstance(value, int) and value <= 10:
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
                    cell.font = Font(color="9C5700")
                elif value == "Not in Top 100":
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    cell.font = Font(color="9C0006")

        ws.row_dimensions[excel_row].height = 20

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
#  STREAMLIT UI
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Rank Checker",
    page_icon="📊",
    layout="wide",
)

st.title("📊 Google Rank Checker")
st.markdown("Upload your keyword sheet, enter your domain and SerpApi key, then run.")

# ── Sidebar inputs ──
with st.sidebar:
    st.header("⚙️ Settings")
    api_key = st.text_input("SerpApi Key", type="password", placeholder="Your SerpApi key")
    domain = st.text_input("Domain to Track", placeholder="e.g. example.com")
    country_name = st.selectbox(
        "Country",
        ["United States", "Canada", "United Kingdom", "Australia", "India", "Pakistan"],
        index=0,
    )
    delay = st.slider("Delay between requests (sec)", 1, 10, 2)

# ── File upload ──
uploaded_file = st.file_uploader(
    "Upload Excel file (.xlsx)",
    type=["xlsx"],
    help="Required columns: Keyword, City, State",
)

if uploaded_file:
    try:
        df = pd.read_excel(uploaded_file)
        df.columns = [c.strip() for c in df.columns]

        # Validate columns
        required_cols = {"Keyword", "City", "State"}
        missing = required_cols - set(df.columns)
        if missing:
            st.error(f"Missing columns: {missing}")
            st.stop()

        st.success(f"✅ Loaded {len(df)} keywords")
        st.dataframe(df, use_container_width=True)

        if st.button("🚀 Start Rank Check", type="primary"):
            if not api_key:
                st.error("Please enter your SerpApi key.")
                st.stop()
            if not domain:
                st.error("Please enter a domain to track.")
                st.stop()

            all_locs = load_locations()
            results = []

            progress_bar = st.progress(0)
            status_text = st.empty()
            results_placeholder = st.empty()

            for i, row in df.iterrows():
                keyword = str(row["Keyword"]).strip()
                city = str(row["City"]).strip()
                state = str(row["State"]).strip()

                status_text.text(f"Checking {i+1}/{len(df)}: {keyword}")

                # ── FIXED: use smart location matcher ──
                targeted_location = find_precise_location(
                    city, state, all_locs, country_name
                )

                organic_rank, found_url = check_organic_rank(
                    keyword, domain, targeted_location, api_key
                )
                maps_rank = check_maps_rank(
                    keyword, domain, targeted_location, api_key
                )

                results.append({
                    "Keyword": keyword,
                    "City": city,
                    "State": state,
                    "Targeted Location": targeted_location,
                    "Organic Rank": organic_rank,
                    "Maps Rank": maps_rank,
                    "Found URL": found_url,
                })

                progress_bar.progress((i + 1) / len(df))

                # Show live results table
                results_df = pd.DataFrame(results)
                results_placeholder.dataframe(results_df, use_container_width=True)

                if i < len(df) - 1:
                    time.sleep(delay)

            status_text.success("✅ Process Completed Successfully!")

            # ── Download button ──
            results_df = pd.DataFrame(results)
            excel_bytes = build_excel_report(results_df)

            st.download_button(
                label="📥 Download Final Report",
                data=excel_bytes,
                file_name="ranking_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    except Exception as e:
        st.error(f"File read error: {e}")
